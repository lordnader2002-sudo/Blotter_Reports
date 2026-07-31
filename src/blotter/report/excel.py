"""Write the rollup to an Excel workbook (primary analyst deliverable)."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ..schema import CATEGORIES

_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_HEADER = Font(bold=True)


def _autosize(ws, max_width: int = 60) -> None:
    for col in ws.columns:
        length = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(length + 2, max_width)


def _naive(df: pd.DataFrame) -> pd.DataFrame:
    """Excel can't store tz-aware datetimes; drop tz for any datetime column."""
    df = df.copy()
    for col in df.columns:
        if isinstance(df[col].dtype, pd.DatetimeTZDtype):
            df[col] = df[col].dt.tz_localize(None)
        elif df[col].dtype == object:
            df[col] = df[col].map(
                lambda v: v.replace(tzinfo=None) if hasattr(v, "tzinfo") and v.tzinfo else v
            )
    return df


def write(rollup, path: str | Path) -> None:
    path = Path(path)
    md = rollup.metadata
    summary = _naive(rollup.summary)
    # `details` is a per-row dict for the dashboard modal — not representable in a cell.
    incidents = _naive(rollup.incidents.drop(columns=["details"], errors="ignore"))
    highlights = _naive(rollup.highlights.drop(columns=["details"], errors="ignore"))

    meta_df = pd.DataFrame(
        [
            {"key": "generated_at", "value": str(md["generated_at"])},
            {"key": "window_days", "value": md["window_days"]},
            {"key": "radius_m", "value": md.get("radius_m", "")},
            {"key": "cutoff", "value": str(md["cutoff"])},
            {"key": "total_incidents", "value": md["total_incidents"]},
            {"key": "violent_incidents", "value": md["violent_incidents"]},
            {"key": "coverage_gaps", "value": ", ".join(md["coverage_gaps"]) or "none"},
        ]
        + [
            {
                "key": f"source:{s.property_id}/{s.name}",
                "value": f"{s.status} fetched={s.fetched_count} "
                f"truncated={s.truncated} {s.error or ''}".strip(),
            }
            for s in md["sources"]
        ]
    )

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        highlights.to_excel(writer, sheet_name="Highlights", index=False)
        incidents.to_excel(writer, sheet_name="All Incidents", index=False)
        meta_df.to_excel(writer, sheet_name="Run Metadata", index=False)

        for name in ("Summary", "Highlights", "All Incidents", "Run Metadata"):
            ws = writer.sheets[name]
            for cell in ws[1]:
                cell.font = _HEADER
            ws.freeze_panes = "A2"
            _autosize(ws)
            if ws.max_row > 1:
                ws.auto_filter.ref = ws.dimensions

        _highlight_summary(writer.sheets["Summary"], summary)


def _highlight_summary(ws, summary: pd.DataFrame) -> None:
    """Red-flag failed/no-coverage malls and malls with any violent incidents."""
    cols = {name: i + 1 for i, name in enumerate(summary.columns)}
    status_col = cols.get("status")
    violent_col = cols.get("VIOLENT" if "VIOLENT" in CATEGORIES else None)
    for row_idx in range(2, ws.max_row + 1):
        if status_col:
            cell = ws.cell(row=row_idx, column=status_col)
            if cell.value in ("FAILED", "NO COVERAGE"):
                cell.fill = _RED
        if violent_col:
            cell = ws.cell(row=row_idx, column=violent_col)
            if isinstance(cell.value, (int, float)) and cell.value > 0:
                cell.fill = _YELLOW
