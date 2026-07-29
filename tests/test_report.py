from datetime import UTC, datetime, timedelta

from openpyxl import load_workbook

from blotter.errors import RunReport, SourceStatus
from blotter.pipeline import RunResult
from blotter.properties import Property
from blotter.report import excel, markdown
from blotter.report.rollup import build_rollup
from blotter.schema import PROPERTY, VIOLENT, NormalizedIncident

NOW = datetime(2026, 6, 19, tzinfo=UTC)


def _result():
    props = {
        "BEVCENTER": Property("BEVCENTER", "Beverly Center", "", "", 34.07533, -118.37738),
        "LENOX": Property("LENOX", "Lenox Square", "", "", 33.8467, -84.3624),
    }
    incidents = [
        NormalizedIncident("BEVCENTER", "s1", "1", NOW - timedelta(days=2), "ROBBERY",
                           VIOLENT, "robbery", "8500 Beverly", 34.0770, -118.3775, 180.0),
        NormalizedIncident("BEVCENTER", "s1", "2", NOW - timedelta(days=5), "BURGLARY",
                           PROPERTY, "burglary", "8400 La Cienega", 34.0760, -118.3780, 320.0),
    ]
    report = RunReport()
    report.sources.append(SourceStatus("BEVCENTER", "s1", "LAPD", "OK", fetched_count=2))
    report.note_coverage_gaps({"LENOX"})
    result = RunResult(incidents, report, NOW, 30, NOW - timedelta(days=30),
                       pilot_ids=["BEVCENTER", "LENOX"])
    return result, props


def test_build_rollup_counts_and_gaps():
    result, props = _result()
    rollup = build_rollup(result, props)
    bev = rollup.summary.set_index("property_id").loc["BEVCENTER"]
    assert bev["total"] == 2
    assert bev[VIOLENT] == 1 and bev[PROPERTY] == 1
    assert bev["nearest_m"] == 180.0
    lenox = rollup.summary.set_index("property_id").loc["LENOX"]
    assert lenox["status"] == "NO COVERAGE"
    assert lenox["total"] == 0
    assert rollup.metadata["violent_incidents"] == 1


def test_markdown_render_contains_sections():
    result, props = _result()
    md = markdown.render(build_rollup(result, props))
    assert "# Mall Blotter Report" in md
    assert "Summary by mall" in md
    assert "No coverage" in md
    assert "Beverly Center" in md


def test_excel_write_creates_sheets(tmp_path):
    result, props = _result()
    out = tmp_path / "report.xlsx"
    excel.write(build_rollup(result, props), out)
    wb = load_workbook(out)
    assert set(wb.sheetnames) == {"Summary", "Highlights", "All Incidents", "Run Metadata"}
    assert wb["All Incidents"].max_row == 3  # header + 2 incidents
